import pandas as pd
import os
import glob
from openpyxl import load_workbook

# Prompt user for the base directory
directory_path = input("Enter the base directory path: ")
# C:\...\...\The University of Southern Mississippi\IR Office - Documents (1)\W Drive\Userfiles\AKale\Resource Allocation Rubric\AY_**_**\FACULTY SUCCESS

##################################################
# Part 1: Load HEGIS Codes and Campus data
##################################################

# Construct the file path for FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")

# Initialize a new Excel workbook
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "HEGIS Data"  # Name of the sheet

# Add the headers for HEGIS Code and Campus in the new workbook
ws.append(["HEGIS Code", "CAMPUS"])

# Construct the file pattern to match the desired Excel file
file_pattern = os.path.join(directory_path, 'INSTRUCTIONAL_FTE_*.xlsx')

# Debug: Print the file pattern being searched
#print(f"Looking for files matching pattern: {file_pattern}")

# Loop through all matching Excel files in the specified directory
matching_files = glob.glob(file_pattern)

# Debug: Print the list of matching files
#print(f"Found files: {matching_files}")

for file_path in matching_files:
    try:
        # Debug: Print the file being processed
        #print(f"Processing file: {file_path}")
        
        # Read the first sheet from the Excel file
        df = pd.read_excel(file_path, sheet_name=0)  # The first sheet
        
        # Debug: Print the first few rows of the DataFrame to verify it's read correctly
        #print(df.head())  # Optional, for debugging

        # Check if the 'HEGIS Code' column exists
        if 'HEGIS Code' in df.columns:
            # Extract the entire 'HEGIS Code' column, including header and values
            hegis_codes = df['HEGIS Code'].dropna().tolist()  # Drop NaN values if any
            
            # Debug: Print the HEGIS codes to verify
            #print(f"HEGIS Codes: {hegis_codes}")
            
            # Append each HEGIS Code and its corresponding CAMPUS rows
            for code in hegis_codes:
                # Append the HEGIS Code and "TOTAL"
                ws.append([code, 'TOTAL'])
                # Append "HGB" and "USMGC" under it
                ws.append([code, 'HBG'])
                ws.append([code, 'USMGC'])
                
        else:
            print(f"'HEGIS Code' column not found in {file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {e}")

# Save the new Excel file
output_path = os.path.join(directory_path, "FS_A.xlsx")
wb.save(output_path)

print(f"Data saved to {output_path}")

##################################################
# Part 2: Load and Merge AR Data Using Wildcard
##################################################

# Step 1: Locate the Applied_Research_AY_* file
ar_file_pattern = os.path.join(directory_path, "Applied_Research_AY_*.xlsx")
ar_files = glob.glob(ar_file_pattern)

if not ar_files:
    print("No Applied_Research_AY_* files found in the directory.")
    exit()
else:
    ar_file_path = ar_files[0]
    print(f"Using file: {ar_file_path}")

# Step 2: Read AR Pivot sheet from the Applied_Research_AY_* file
try:
    ar_pivot_data = pd.read_excel(ar_file_path, sheet_name="AR Pivot")
    print("AR Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading AR Pivot sheet: {e}")
    exit()

# Step 3: Ensure required columns are present in AR Pivot
required_columns = ['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'score']
if not all(col in ar_pivot_data.columns for col in required_columns):
    print(f"Missing required columns in AR Pivot sheet. Expected: {required_columns}")
    exit()

# Normalize data for matching (strip spaces and make lowercase)
ar_pivot_data['HEGIS Code'] = ar_pivot_data['HEGIS Code'].astype(str).str.strip()
ar_pivot_data['Home Campus/Teaching Site (Most Recent)'] = ar_pivot_data['Home Campus/Teaching Site (Most Recent)'].str.strip().str.lower()

# Step 4: Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 5: Merge FS_A data with AR Pivot data on HEGIS Code and CAMPUS
mapped_data = fs_a_data.merge(
    ar_pivot_data[['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)']
)

# Step 5: Handle missing scores for HBG and USMGC (set them to 0.0)
mapped_data['score'] = mapped_data['score'].fillna(0.0)  # Fix: Assign to the column instead of inplace

# Step 6: Group and Sum Scores for 'HBG' and 'USMGC'
total_scores = mapped_data[mapped_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['score'].sum().reset_index()

# Step 7: Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_data = mapped_data.merge(total_scores, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Step 8: Update the TOTAL column for each HEGIS Code and Campus
mapped_data['TOTAL'] = mapped_data.apply(
    lambda row: row['score_total'] if row['CAMPUS'] == 'total' else row['score'], axis=1
)

# Clean up the temporary column
mapped_data.drop(columns=['score_total'], inplace=True)

# Step 9: Drop the 'score' column
mapped_data.drop(columns=['score'], inplace=True)

# Step 10: Write the updated data to the Excel file

# Load the workbook and sheet
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Ensure the column headers are present (add them to the first row if they don't exist)
headers = ['HEGIS Code', 'CAMPUS', 'AR TOTAL']
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header)

# Ensure that we start writing from the second row (skipping the header)
for index, row in mapped_data.iterrows():
    # Excel rows are 1-based, while pandas rows are 0-based
    excel_row = index + 2  # Add 2 to skip the header row

    # Write the 'HEGIS Code', 'CAMPUS', and 'TOTAL' values into the appropriate columns
    ws.cell(row=excel_row, column=1, value=row['HEGIS Code'])  # HEGIS Code
    ws.cell(row=excel_row, column=2, value=row['CAMPUS'])      # CAMPUS
    ws.cell(row=excel_row, column=3, value=row['TOTAL'])    # TOTAL

# Step 11: Save the updated workbook
wb.save(fs_a_path)

print(f"FS_A.xlsx has been successfully updated with AR totals.")

###################################################
# Part 3: Load and Merge Awards Data Using Wildcard
###################################################

# Read the Awards Pivot sheet from the Awards_AY_* file
awards_file_pattern = os.path.join(directory_path, "Awards_AY_*.xlsx")
awards_files = glob.glob(awards_file_pattern)

if not awards_files:
    print("No Awards_AY_* files found in the directory.")
    exit()
else:
    awards_file_path = awards_files[0]
    print(f"Using file: {awards_file_path}")

# Read Awards Pivot sheet from the Awards_AY_* file
try:
    awards_pivot_data = pd.read_excel(awards_file_path, sheet_name="Awards Pivot")
    print("Awards Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading Awards Pivot sheet: {e}")
    exit()

# Ensure required columns are present in Awards Pivot
required_columns = ['ID_String_Multiplied', 'Location', 'HEGIS Code']
if not all(col in awards_pivot_data.columns for col in required_columns):
    print(f"Missing required columns in Awards Pivot sheet. Expected: {required_columns}")
    exit()

# Normalize data for matching (strip spaces and make lowercase)
awards_pivot_data['HEGIS Code'] = awards_pivot_data['HEGIS Code'].astype(str).str.strip()
awards_pivot_data['Location'] = awards_pivot_data['Location'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Merge FS_A data with Awards Pivot data on HEGIS Code and Location
mapped_awards_data = fs_a_data.merge(
    awards_pivot_data[['HEGIS Code', 'Location', 'ID_String_Multiplied']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Location']
)

# Handle missing ID_String_Multiplied (set to 0.0 for missing)
mapped_awards_data['ID_String_Multiplied'] = mapped_awards_data['ID_String_Multiplied'].fillna(0.0)

# Calculate the AWARDS TOTAL by summing HBG and USMGC
# Create a new dataframe for HBG and USMGC sums
total_awards = mapped_awards_data[mapped_awards_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['ID_String_Multiplied'].sum().reset_index()

# Merge the total awards with the original data to add the 'AWARDS TOTAL' column for each HEGIS Code
mapped_awards_data = mapped_awards_data.merge(total_awards, on='HEGIS Code', how='left', suffixes=('', '_awards_total'))

# Update the AWARDS TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_awards_data['AWARDS TOTAL'] = mapped_awards_data.apply(
    lambda row: row['ID_String_Multiplied_awards_total'] if row['CAMPUS'] == 'total' else row['ID_String_Multiplied'], axis=1
)

# Clean up the temporary column
mapped_awards_data = mapped_awards_data.drop(columns=['ID_String_Multiplied_awards_total'])

# Step 3: Update FS_A.xlsx with the calculated AWARDS TOTAL
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column after AR TOTAL
header_row = 1
next_column = ws.max_column + 1

# Add 'AWARDS TOTAL' header
ws.cell(row=header_row, column=next_column, value='AWARDS TOTAL')

# Update the AWARDS TOTAL values in the next column
for index, row in mapped_awards_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['AWARDS TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with AWARDS TOTAL.")

############################################################
# Part 4: Load and Merge Creative Works Data Using Wildcard
############################################################

# Read the Creative Works Pivot sheet from the Creative_Works_AY_* file
cw_file_pattern = os.path.join(directory_path, "Creative_Works_AY_*.xlsx")
cw_files = glob.glob(cw_file_pattern)

if not cw_files:
    print("No Creative_Works_AY_* files found in the directory.")
    exit()
else:
    cw_file_path = cw_files[0]
    print(f"Using file: {cw_file_path}")

# Read CW Pivot sheet from the Creative_Works_AY_* file
try:
    cw_pivot_data = pd.read_excel(cw_file_path, sheet_name="CW Pivot")
    print("CW Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading CW Pivot sheet: {e}")
    exit()

# Ensure required columns are present in CW Pivot
required_columns = ['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'score']
if not all(col in cw_pivot_data.columns for col in required_columns):
    print(f"Missing required columns in CW Pivot sheet. Expected: {required_columns}")
    exit()

# Normalize data for matching (strip spaces and make lowercase)
cw_pivot_data['HEGIS Code'] = cw_pivot_data['HEGIS Code'].astype(str).str.strip()
cw_pivot_data['Home Campus/Teaching Site (Most Recent)'] = cw_pivot_data['Home Campus/Teaching Site (Most Recent)'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Merge FS_A data with CW Pivot data on HEGIS Code and CAMPUS
mapped_cw_data = fs_a_data.merge(
    cw_pivot_data[['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)']
)

# Handle missing scores (set to 0.0 for missing)
mapped_cw_data['score'] = mapped_cw_data['score'].fillna(0.0)

# Group by HEGIS Code and CAMPUS to calculate the total score for HBG and USMGC
# We will use this to calculate the 'TOTAL' for each HEGIS Code
total_scores_cw = mapped_cw_data[mapped_cw_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_cw_data = mapped_cw_data.merge(total_scores_cw, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_cw_data['TOTAL'] = mapped_cw_data.apply(
    lambda row: row['score_total'] if row['CAMPUS'] == 'total' else row['score'], axis=1
)

# Clean up the temporary column (without using inplace)
mapped_cw_data = mapped_cw_data.drop(columns=['score_total'])

# Debug: Display the updated data to ensure it meets the expected output
#print(mapped_cw_data)

# Step 3: Update FS_A.xlsx with the calculated CW SCORE and TOTAL
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'CW SCORE' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='CW SCORE')

# Update the CW SCORE values in the next column
for index, row in mapped_cw_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['score'])  # This adds CW SCORE values

# Add 'CW TOTAL' next to CW SCORE
total_column = next_column + 1
ws.cell(row=header_row, column=total_column, value='CW TOTAL')

# Update the CW TOTAL values
for index, row in mapped_cw_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=total_column, value=row['TOTAL'])  # This adds CW TOTAL values

# Remove the CW SCORE column by deleting the column where CW SCORE is located (next_column)
ws.delete_cols(next_column)

# Save the updated FS_A.xlsx file with only CW TOTAL
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with CW TOTAL.")

#####################################################
# Part 5: Load and Merge Grants Data Using Wildcard
#####################################################

# Step 1: Load the Grants_AY_* file
grants_file_pattern = os.path.join(directory_path, "Grants_AY_*.xlsx")
grants_files = glob.glob(grants_file_pattern)

if not grants_files:
    print("No Grants_AY_* files found in the directory.")
    exit()
else:
    grants_file_path = grants_files[0]
    print(f"Using file: {grants_file_path}")

# Read the Grants Pivot sheet
try:
    grants_pivot_data = pd.read_excel(grants_file_path, sheet_name="GN Pivot")
    print("Grants Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading Grants Pivot sheet: {e}")
    exit()

# Check for required columns and map them correctly
if not all(col in grants_pivot_data.columns for col in ['HEGIS_Code', 'Location']):
    print("Missing required columns in Grants Pivot sheet. Expected: ['HEGIS_Code', 'Location']")
    exit()

# Map the 'ID x 1.1' column to 'score' for consistency with the previous process
grants_pivot_data['score'] = grants_pivot_data['ID x 1.1']  # Use ID x 1.1 as the score column

# Normalize data for matching (strip spaces and make lowercase)
grants_pivot_data['HEGIS_Code'] = grants_pivot_data['HEGIS_Code'].astype(str).str.strip()
grants_pivot_data['Location'] = grants_pivot_data['Location'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 2: Merge FS_A data with Grants Pivot data on HEGIS Code and CAMPUS
mapped_grants_data = fs_a_data.merge(
    grants_pivot_data[['HEGIS_Code', 'Location', 'score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS_Code', 'Location']
)

# Handle missing scores (set to 0.0 for missing)
mapped_grants_data['score'] = mapped_grants_data['score'].fillna(0.0)

# Step 3: Group by HEGIS Code and CAMPUS to calculate the total score for HBG and USMGC
total_scores_grants = mapped_grants_data[mapped_grants_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_grants_data = mapped_grants_data.merge(total_scores_grants, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_grants_data['TOTAL'] = mapped_grants_data.apply(
    lambda row: row['score_total'] if row['CAMPUS'] == 'total' else row['score'], axis=1
)

# Clean up the temporary column
mapped_grants_data = mapped_grants_data.drop(columns=['score_total'])

# Debug: Display the updated data to ensure it meets the expected output
#print(mapped_grants_data)

# Step 4: Update FS_A.xlsx with the calculated GRANTS TOTAL (only)
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'GRANTS TOTAL' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='GRANTS TOTAL')

# Update the GRANTS TOTAL values
for index, row in mapped_grants_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with GRANTS TOTAL.")

################################################
# Part 6: Load and Merge IP Data Using Wildcard
################################################

# Step 1: Load the IP_AY_* file
ip_file_pattern = os.path.join(directory_path, "IP_AY_*.xlsx")
ip_files = glob.glob(ip_file_pattern)

if not ip_files:
    print("No IP_AY_* files found in the directory.")
    exit()
else:
    ip_file_path = ip_files[0]
    print(f"Using file: {ip_file_path}")

# Read the IP Pivot sheet
try:
    ip_pivot_data = pd.read_excel(ip_file_path, sheet_name="IP Pivot")
    print("IP Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading IP Pivot sheet: {e}")
    exit()

# Check for required columns and map them correctly
if not all(col in ip_pivot_data.columns for col in ['HEGIS Code', 'Location', 'Score']):
    print("Missing required columns in IP Pivot sheet. Expected: ['HEGIS Code', 'Location', 'Score']")
    exit()

# Normalize data for matching (strip spaces and make lowercase)
ip_pivot_data['HEGIS Code'] = ip_pivot_data['HEGIS Code'].astype(str).str.strip()
ip_pivot_data['Location'] = ip_pivot_data['Location'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 2: Merge FS_A data with IP Pivot data on HEGIS Code and CAMPUS
mapped_ip_data = fs_a_data.merge(
    ip_pivot_data[['HEGIS Code', 'Location', 'Score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Location']
)

# Handle missing scores (set to 0.0 for missing)
mapped_ip_data['Score'] = mapped_ip_data['Score'].fillna(0.0)

# Step 3: Group by HEGIS Code and CAMPUS to calculate the total score for each
total_scores_ip = mapped_ip_data[mapped_ip_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['Score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_ip_data = mapped_ip_data.merge(total_scores_ip, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_ip_data['TOTAL'] = mapped_ip_data.apply(
    lambda row: row['Score_total'] if row['CAMPUS'] == 'total' else row['Score'], axis=1
)

# Clean up the temporary column
mapped_ip_data = mapped_ip_data.drop(columns=['Score_total'])

# Debug: Display the updated data to ensure it meets the expected output
#print(mapped_ip_data)

# Step 4: Update FS_A.xlsx with the calculated IP TOTAL (only)
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'IP TOTAL' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='IP TOTAL')

# Update the IP TOTAL values
for index, row in mapped_ip_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with IP TOTAL.")

############################################################
# PART 7: Process Publications Files
###########################################################
print("Processing Publications_AY_**_**_updated files...")

# Step 1: List all files in the directory for debugging
all_files = os.listdir(directory_path)
print(f"Files in the directory: {all_files}")

# Step 2: Load the Publications_AY_*_updated.xlsx file
pub_file_pattern = os.path.join(directory_path, "Publications_AY_*_*_updated.xlsx")
pub_files = glob.glob(pub_file_pattern)
print(f"Matched files: {pub_files}")

if not pub_files:
    print("No Publications_AY_*_updated files found in the directory.")
    exit()
else:
    pub_file_path = pub_files[0]
    print(f"Using file: {pub_file_path}")

# Read the Pivot_Table sheet
try:
    pivot_table_data = pd.read_excel(pub_file_path, sheet_name="Pivot_Table")
    print("Pivot_Table sheet loaded successfully.")
except Exception as e:
    print(f"Error loading Pivot_Table sheet: {e}")
    exit()

# Check for required columns and map them correctly
if not all(col in pivot_table_data.columns for col in ['HEGIS Code', 'Location_y', 'adjusted_total_score']):
    print("Missing required columns in Pivot_Table sheet. Expected: ['HEGIS Code', 'Locatio_y', 'adjusted_total_score']")
    exit()

# Normalize data for matching (strip spaces and make lowercase)
pivot_table_data['HEGIS Code'] = pivot_table_data['HEGIS Code'].astype(str).str.strip()
pivot_table_data['Location_y'] = pivot_table_data['Location_y'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
try:
    fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")
    print("FS_A HEGIS Data sheet loaded successfully.")
except Exception as e:
    print(f"Error loading FS_A.xlsx HEGIS Data sheet: {e}")
    exit()

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 2: Merge FS_A data with Pivot_Table data on HEGIS Code and CAMPUS
mapped_pub_data = fs_a_data.merge(
    pivot_table_data[['HEGIS Code', 'Location_y', 'adjusted_total_score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Location_y']
)

# Handle missing scores (set to 0.0 for missing)
mapped_pub_data['adjusted_total_score'] = mapped_pub_data['adjusted_total_score'].fillna(0.0)

# Step 3: Group by HEGIS Code and CAMPUS to calculate the total score for each
total_scores_pub = mapped_pub_data[mapped_pub_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['adjusted_total_score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_pub_data = mapped_pub_data.merge(total_scores_pub, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_pub_data['TOTAL'] = mapped_pub_data.apply(
    lambda row: row['adjusted_total_score_total'] if row['CAMPUS'] == 'total' else row['adjusted_total_score'], axis=1
)

# Clean up the temporary column
mapped_pub_data = mapped_pub_data.drop(columns=['adjusted_total_score_total'])

# Step 4: Update FS_A.xlsx with the calculated Publications TOTAL (only)
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'PUBLICATIONS TOTAL' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='PUBLICATIONS TOTAL')

# Update the PUBLICATIONS TOTAL values
for index, row in mapped_pub_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with PUBLICATIONS TOTAL.")

###########################################################
# Part 8: Load and Merge Presintations Data Using Wildcard
###########################################################

# Step 1: Load the Presentations_AY_* file
presentations_file_pattern = os.path.join(directory_path, "Presentations_AY_*.xlsx")
presentations_files = glob.glob(presentations_file_pattern)

if not presentations_files:
    print("No Presentations_AY_* files found in the directory.")
    exit()
else:
    presentations_file_path = presentations_files[0]
    print(f"Using file: {presentations_file_path}")

# Read the Presentations Pivot sheet
try:
    presentations_pivot_data = pd.read_excel(presentations_file_path, sheet_name="Presentations Pivot")
    print("Presentations Pivot sheet loaded successfully.")
except Exception as e:
    print(f"Error loading Presentations Pivot sheet: {e}")
    exit()

# Check for required columns and map them correctly
if not all(col in presentations_pivot_data.columns for col in ['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'INVACC', 'INVACC_Updated']):
    print("Missing required columns in Presentations Pivot sheet. Expected: ['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'INVACC', 'INVACC_Updated']")
    exit()

# Map the 'INVACC_Updated' column to 'score' for consistency with the previous process
presentations_pivot_data['score'] = presentations_pivot_data['INVACC_Updated']  # Use INVACC_Updated as the score column

# Normalize data for matching (strip spaces and make lowercase)
presentations_pivot_data['HEGIS Code'] = presentations_pivot_data['HEGIS Code'].astype(str).str.strip()
presentations_pivot_data['Home Campus/Teaching Site (Most Recent)'] = presentations_pivot_data['Home Campus/Teaching Site (Most Recent)'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 2: Merge FS_A data with Presentations Pivot data on HEGIS Code and Home Campus
mapped_presentations_data = fs_a_data.merge(
    presentations_pivot_data[['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)']
)

# Handle missing scores (set to 0.0 for missing)
mapped_presentations_data['score'] = mapped_presentations_data['score'].fillna(0.0)

# Step 3: Group by HEGIS Code and CAMPUS to calculate the total score for HBG and USMGC
total_scores_presentations = mapped_presentations_data[mapped_presentations_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_presentations_data = mapped_presentations_data.merge(total_scores_presentations, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_presentations_data['PRESENTATIONS TOTAL'] = mapped_presentations_data.apply(
    lambda row: row['score_total'] if row['CAMPUS'] == 'total' else row['score'], axis=1
)

# Clean up the temporary column
mapped_presentations_data = mapped_presentations_data.drop(columns=['score_total'])

# Debug: Display the updated data to ensure it meets the expected output
#print(mapped_presentations_data)

# Step 4: Update FS_A.xlsx with the calculated PRESENTATIONS TOTAL (only)
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the column that contains 'Presentations TOTAL' and remove it if it exists
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Presentations TOTAL":
        ws.delete_cols(col)
        #print("Found and removed 'Presentations TOTAL' column.")

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'PRESENTATIONS TOTAL' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='PRESENTATIONS TOTAL')

# Update the PRESENTATIONS TOTAL values
for index, row in mapped_presentations_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['PRESENTATIONS TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

print(f"FS_A.xlsx has been updated with PRESENTATIONS TOTAL.")

##########################################################
# Part 9: Load and Merge Publications Data Using Wildcard
##########################################################

# Step 1: Load the Publications_AY_* file
publications_file_pattern = os.path.join(directory_path, "Publications_AY_*_*_updated.xlsx")
publications_files = glob.glob(publications_file_pattern)

if not publications_files:
    print("No Publications_AY_*_*_updated files found in the directory.")
    exit()
else:
    publications_file_path = publications_files[0]
    print(f"Using file: {publications_file_path}")

# Read the Pivot Table sheet
try:
    publications_pivot_data = pd.read_excel(publications_file_path, sheet_name="Pivot_Table")
    print("Pivot Table sheet loaded successfully.")
except Exception as e:
    print(f"Error loading Pivot Table sheet: {e}")
    exit()

# Check for required columns and map them correctly
if not all(col in publications_pivot_data.columns for col in ['HEGIS Code', 'Location_y', 'total_score', 'adjusted_total_score']):
    print("Missing required columns in Pivot Table sheet. Expected: ['HEGIS Code', 'Location_y', 'total_score', 'adjusted_total_score']")
    exit()

# Map the 'adjusted_total_score' column to 'score' for consistency with the previous process
publications_pivot_data['score'] = publications_pivot_data['adjusted_total_score']  # Use adjusted_total_score as the score column

# Normalize data for matching (strip spaces and make lowercase)
publications_pivot_data['HEGIS Code'] = publications_pivot_data['HEGIS Code'].astype(str).str.strip()
publications_pivot_data['Location_y'] = publications_pivot_data['Location_y'].str.strip().str.lower()

# Load FS_A.xlsx
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Normalize FS_A data for matching
fs_a_data['HEGIS Code'] = fs_a_data['HEGIS Code'].astype(str).str.strip()
fs_a_data['CAMPUS'] = fs_a_data['CAMPUS'].str.strip().str.lower()

# Step 2: Merge FS_A data with Publications Pivot data on HEGIS Code and Location_master
mapped_publications_data = fs_a_data.merge(
    publications_pivot_data[['HEGIS Code', 'Location_y', 'score']],
    how="left",
    left_on=['HEGIS Code', 'CAMPUS'],
    right_on=['HEGIS Code', 'Location_y']
)

# Handle missing scores (set to 0.0 for missing)
mapped_publications_data['score'] = mapped_publications_data['score'].fillna(0.0)

# Step 3: Group by HEGIS Code and CAMPUS to calculate the total score for HBG and USMGC
total_scores_publications = mapped_publications_data[mapped_publications_data['CAMPUS'].isin(['hbg', 'usmgc'])] \
    .groupby('HEGIS Code')['score'].sum().reset_index()

# Merge the total scores with the original data to add the 'TOTAL' column for each HEGIS Code
mapped_publications_data = mapped_publications_data.merge(total_scores_publications, on='HEGIS Code', how='left', suffixes=('', '_total'))

# Update the TOTAL column: only the row with 'TOTAL' as the campus will have the sum
mapped_publications_data['PUBLICATIONS TOTAL'] = mapped_publications_data.apply(
    lambda row: row['score_total'] if row['CAMPUS'] == 'total' else row['score'], axis=1
)

# Clean up the temporary column
mapped_publications_data = mapped_publications_data.drop(columns=['score_total'])

# Debug: Display the updated data to ensure it meets the expected output
#print(mapped_publications_data)

# Step 4: Update FS_A.xlsx with the calculated PUBLICATIONS TOTAL (only)
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the column that contains 'Publications TOTAL' and remove it if it exists
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Publications TOTAL":
        ws.delete_cols(col)
        print("Found and removed 'Publications TOTAL' column.")

# Find the next available column after AWARDS TOTAL (already in the workbook)
header_row = 1
next_column = ws.max_column + 1

# Add 'PUBLICATIONS TOTAL' header next to AWARDS TOTAL
ws.cell(row=header_row, column=next_column, value='PUBLICATIONS TOTAL')

# Update the PUBLICATIONS TOTAL values
for index, row in mapped_publications_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['PUBLICATIONS TOTAL'])

# Save the updated FS_A.xlsx file
wb.save(fs_a_path)

#print(f"FS_A.xlsx has been updated with PUBLICATIONS TOTAL.")

##########################################################
# Part 10: Define OUTPUT Folder
##########################################################

# Define the output folder path, which is one level up from the directory path
output_folder = os.path.join(os.path.dirname(directory_path), "OUTPUT")  # Parent directory + OUTPUT

# Ensure the output directory exists
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

##########################################################
# Part 11: Adding all columns together
##########################################################

# Load FS_A.xlsx and the "HEGIS Data" sheet
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Step 1: Sum all numeric rows for each row
numeric_columns = fs_a_data.select_dtypes(include='number').columns  # Only numeric columns

# Create a new column with the sum of all numeric columns for each row
fs_a_data['Total Row Sum'] = fs_a_data[numeric_columns].sum(axis=1)

# Step 2: Multiply the "Total Row Sum" by 0.175
fs_a_data['Total Row Sum'] *= 0.175

# Step 3: Update the FS_A.xlsx with the new "Total Row Sum" column
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Find the next available column (assuming the Total Row Sum will be added to the last column)
header_row = 1
next_column = ws.max_column + 1

# Add the "Total Row Sum" header
ws.cell(row=header_row, column=next_column, value='Total Row Sum')

# Update the "Total Row Sum" values in the new column
for index, row in fs_a_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['Total Row Sum'])

# Save the updated FS_A.xlsx file in the OUTPUT folder
output_fs_a_path = os.path.join(output_folder, "FS_A_updated.xlsx")
wb.save(output_fs_a_path)

# Also save the updated file to the original location
wb.save(fs_a_path)

#print(f"FS_A.xlsx has been updated with the Total Row Sum multiplied by 0.175 and saved to {output_fs_a_path}.")
#print(f"The original FS_A.xlsx has also been updated and saved back to {fs_a_path}.")

##########################################################
# Part 12: Flattening Total Row Sum by HEGIS and Campus
##########################################################

# Load FS_A.xlsx and the "HEGIS Data" sheet
fs_a_path = os.path.join(directory_path, "FS_A.xlsx")
fs_a_data = pd.read_excel(fs_a_path, sheet_name="HEGIS Data")

# Columns to sum for the Total Row Sum
sum_columns = ['AR TOTAL', 'AWARDS TOTAL', 'CW TOTAL', 'GRANTS TOTAL', 'IP TOTAL', 'PRESENTATIONS TOTAL', 'PUBLICATIONS TOTAL']

# Step 1: Sum specific numeric columns for each row
fs_a_data['Total Row Sum'] = fs_a_data[sum_columns].sum(axis=1)

# Step 2: Multiply the "Total Row Sum" by 0.175 and round to 2 decimal places
fs_a_data['Total Row Sum'] = (fs_a_data['Total Row Sum'] * 0.175).round(2)

# Step 3: Update the FS_A.xlsx with the new "Total Row Sum" column
wb = load_workbook(fs_a_path)
ws = wb["HEGIS Data"]

# Check if the "Total Row Sum" already exists in the sheet
header_row = 1
existing_columns = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

# Find the next available column if "Total Row Sum" doesn't already exist
if 'Total Row Sum' not in existing_columns:
    next_column = ws.max_column + 1
    # Add the "Total Row Sum" header
    ws.cell(row=header_row, column=next_column, value='Total Row Sum')
else:
    next_column = existing_columns.index('Total Row Sum') + 1  # Update the column if it exists

# Update the "Total Row Sum" values in the new or existing column
for index, row in fs_a_data.iterrows():
    excel_row = index + 2  # Add 2 to skip the header row
    ws.cell(row=excel_row, column=next_column, value=row['Total Row Sum'])

# Save the updated FS_A.xlsx file in the OUTPUT folder
wb.save(output_fs_a_path)

# Also save the updated file to the original location
wb.save(fs_a_path)

#print(f"FS_A.xlsx has been updated with the Total Row Sum multiplied by 0.175 and saved to {output_fs_a_path}.")
#print(f"The original FS_A.xlsx has also been updated and saved back to {fs_a_path}.")

# Step 1: Pivot the data to flatten
pivot_data = fs_a_data.pivot_table(index='HEGIS Code', columns='CAMPUS', values='Total Row Sum', aggfunc='sum', fill_value=0)

# Step 2: Create a new sheet for the flattened data
wb = load_workbook(output_fs_a_path)

# Remove the 'Flattened Data' sheet if it already exists
if 'Flattened Data' in wb.sheetnames:
    sheet_to_remove = wb['Flattened Data']
    wb.remove(sheet_to_remove)

# Add a new 'Flattened Data' sheet
flattened_ws = wb.create_sheet(title='Flattened Data')

# Step 3: Write the flattened data into the new sheet
# Write headers
flattened_ws.append(['HEGIS Code'] + list(pivot_data.columns))

# Write data rows
for hegis_code, row_values in pivot_data.iterrows():
    flattened_ws.append([hegis_code] + list(row_values))

# Save the workbook with the flattened data to the output folder
wb.save(output_fs_a_path)

# Final message indicating completion
print(f"FS_A.xlsx has been updated with a new sheet 'Flattened Data' and saved to {output_fs_a_path}.")
print(f"The flattened data includes HEGIS Code by Campus with the Total Row Sum calculated.")