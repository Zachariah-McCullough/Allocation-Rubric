# THE OUTFILE HIP_B WILL NOT WRITE THE PATHWAY YOU GIVE IT, IT WILL ONLY WRITE TO THE OUTPUT FOLDER

import pandas as pd
import os
import glob
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Prompt user for the base directory
directory_path = input("Enter the base directory path: ")
# C:\...\...\The University of Southern Mississippi\IR Office - Documents (1)\W Drive\Userfiles\AKale\Resource Allocation Rubric\AY_**_**\FACULTY SUCCESS
# The first found file is read, specifically the High Impact Practices sheet.
# It filters the rows for records where TYPE is either "Mentored Student Creative Activity" or "Mentored Student Publication" and where COMPSTAGE is one of "Completed", "In-Process", or "Published".
# Create Pivot Table: A pivot table is created to count occurrences of ID_String by HEGIS Code and Home Campus/Teaching Site (Most Recent), and it is renamed to Count.
# Write Pivot Table to File: The pivot table is saved to the Excel file under a sheet named Count.
# Stage 2: Processing High_Impact_Practices_Scheduled_Learning Files
    # Locate Files: It searches for files matching the pattern High_Impact_Practices_Scheduled_Learning_AY_*.xlsx.
    # Create Pivot Table: It creates a pivot table for the Scheduled Learning sheet, summing IMPACT_ASL by HEGIS Code and Home Campus/Teaching Site (Most Recent), and renames it to Count.
    # Add Additional Data: A new column, ASL, is added, calculated by multiplying Count by 6.
    # Write Pivot Table to File: The pivot table is saved to the Excel file under a sheet named ASL.
# Stage 3: Additional HIP Pivot Table Creation for Scheduled Learning
    # Create HIP Pivot Table: A new pivot table is created for the same file (Scheduled Learning), counting IMPACT by HEGIS Code and Home Campus/Teaching Site (Most Recent), and the count is multiplied by 2 to create the HIP column.
    # Write to HIP Sheet: The pivot table is appended to a sheet named HIP. If the HIP sheet exists, the table is added below the last row. Otherwise, it's created from scratch.
    # Stage 4: Combining Data and Final Data Processing
# The code extracts and merges data from the HIP, ASL, and Count sheets across multiple files.
# The data is merged on HEGIS Code and Home Campus/Teaching Site (Most Recent).
# Calculate Additional Columns:
    # It calculates a sum of columns (HIP.1, ASL, Count) and computes a Weighted Sum based on a weight of 0.175.
    # It ensures that "HBG" (Hattiesburg) and "USMGC" (Gulf Coast) rows exist in the data for each HEGIS Code. If any of these rows are missing, they are added with zero values.
    # It also calculates a TOTAL row for each HEGIS Code.
# Write Merged Data to File:
    # The final merged data is saved to a new Excel file (HIP_B.xlsx) in the OUTPUT folder.
    # The merged data is pivoted again to create a more concise view, with HEGIS Code as the index and columns for HIP.1, ASL, Count, sum, and Weighted Sum, split by Home Campus/Teaching Site (Most Recent).
    # The multi-level columns are flattened.
    # Write Flattened Data: The flattened data is written to a new sheet, Flattened Data, in the same output file.
# Output Files
    # The processed and merged data is saved in the OUTPUT folder as HIP_B.xlsx.
# The file includes:
    # A Count sheet (from Directed Service Learning).
    # An ASL sheet (from Scheduled Learning).
    # A HIP sheet (with additional HIP-related data).
    # A Merged Data sheet (with the combined data from all sources).
    # A Flattened Data sheet (with the data in a simplified format).
# The code includes error handling for situations where the files or sheets are missing, ensuring that any issues encountered during the processing are reported.

########################################################### PART 1: Process High_Impact_Practices_Directed_Service_Learning #########################################################
print("Stage 1: Processing High_Impact_Practices_Directed_Service_Learning files...")

# Use glob to find the file in the specified directory
file_path_directed = glob.glob(os.path.join(directory_path, "High_Impact_Practices_Directed_Service_Learning_AY_*.xlsx"))

# Ensure at least one file is found
if not file_path_directed:
    print("No file found with the specified pattern for Directed Service Learning.")
else:
    # Load the first matching file, specifically the 'High Impact Practices' sheet
    df_directed = pd.read_excel(file_path_directed[0], sheet_name='High Impact Practices')

    # Apply the filters for TYPE and COMPSTAGE
    filtered_df_directed = df_directed[
        (df_directed['TYPE'].isin(['Mentored Student Creative Activity', 'Mentored Student Publication'])) &
        (df_directed['COMPSTAGE'].isin(['Completed', 'In-Process', 'Published']))
    ]

    # Create the pivot table
    pivot_table_directed = filtered_df_directed.pivot_table(
        values='ID_String',
        index=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)'],
        aggfunc='count'
    ).rename(columns={'ID_String': 'Count'})

    # Print the pivot table for Directed Service Learning
    #print("\nPivot Table for Directed Service Learning:")
    #print(pivot_table_directed)

    # Write the pivot table to a new sheet called "ASL" in the same file, replacing it if it already exists
    try:
        with pd.ExcelWriter(file_path_directed[0], engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
            pivot_table_directed.to_excel(writer, sheet_name='Count')
        print("Stage 1 Complete: Pivot table for Directed Service Learning written to the 'Count' sheet.")
    except Exception as e:
        print(f"Error writing Directed Service Learning pivot table: {e}")

########################################################### PART 2: Process High_Impact_Practices_Scheduled_Learning ################################################################
print("Stage 2: Processing High_Impact_Practices_Scheduled_Learning files...")

# Use glob to find the file for Scheduled Learning in the specified directory
file_path_scheduled = glob.glob(os.path.join(directory_path, "High_Impact_Practices_Scheduled_Learning_AY_*.xlsx"))

# Ensure at least one file is found
if not file_path_scheduled:
    print("No file found with the specified pattern for Scheduled Learning.")
else:
    # Load the first matching file, specifically the 'Scheduled Learning' sheet
    df_scheduled = pd.read_excel(file_path_scheduled[0], sheet_name='Scheduled Learning')

    # Create the pivot table with 'Count' as the value column
    pivot_table_scheduled = df_scheduled.pivot_table(
        values='IMPACT_ASL',
        index=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)'],
        aggfunc='sum'
    ).rename(columns={'IMPACT_ASL': 'Count'})

    # Add a new column 'ASL' that is Count multiplied by 6
    pivot_table_scheduled['ASL'] = pivot_table_scheduled['Count'] * 6

    # Print the pivot table for Scheduled Learning
    #print("\nPivot Table for Scheduled Learning (with 'ASL' column):")
    #print(pivot_table_scheduled)

    # Write the pivot table to the 'ASL' sheet in the same file, replacing it if the sheet already exists
    try:
        with pd.ExcelWriter(file_path_scheduled[0], engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
            pivot_table_scheduled.to_excel(writer, sheet_name='ASL')
        print("Stage 2 Complete: Pivot table for Scheduled Learning written to the 'ASL' sheet with 'Count' and 'ASL' columns.")
    except Exception as e:
        print(f"Error writing Scheduled Learning pivot table: {e}")

########################################################### PART 3: Additional Pivot Table for High_Impact_Practices_Scheduled_Learning #############################################
print("Stage 3: Creating additional HIP pivot table for High_Impact_Practices_Scheduled_Learning file...")

# Ensure we have already loaded file_path_scheduled in previous part
if not file_path_scheduled:
    print("No file found with the specified pattern for Scheduled Learning.")
else:
    # Create the additional HIP pivot table
    pivot_table_hip = df_scheduled.pivot_table(
        values='IMPACT',
        index=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)'],
        aggfunc='count'
    ).rename(columns={'IMPACT': 'Count'})

    # Add a new column 'HIP' that is Count multiplied by 2
    pivot_table_hip['HIP'] = pivot_table_hip['Count'] * 2

    # Print the pivot table for HIP
    #print("\nPivot Table for HIP (High Impact Practices):")
    #print(pivot_table_hip)

    # Append this pivot table to the 'HIP' sheet in the same Scheduled Learning Excel file
    try:
        with pd.ExcelWriter(file_path_scheduled[0], engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
            # Check if the 'HIP' sheet exists
            if 'HIP' in writer.sheets:
                start_row = writer.sheets['HIP'].max_row  # Find the last row in 'HIP' sheet
            else:
                start_row = 0  # If 'HIP' does not exist, write at the start

            # Write the HIP pivot table to the 'HIP' sheet
            pivot_table_hip.to_excel(writer, sheet_name='HIP', startrow=start_row, header=(start_row == 0))

            # If we're starting from the top, write header for HIP
            if start_row == 0:  # If we're starting from the top, we need to ensure the header is correctly defined
                worksheet = writer.sheets['HIP']
                worksheet.cell(row=1, column=3, value='HIP')  # Writing header for HIP

    except Exception as e:
        print(f"Error writing HIP pivot table: {str(e)}")  # Print the error message for more context

    print("Stage 3 Complete: Additional HIP pivot table written to the 'HIP' sheet.")

########################################################### PART 4: Combining all data points and flattening #########################################################################
print("Stage 4: Extracting the HIP sheet and merging ASL and Count columns from the relevant sheets...")

# Load the necessary files and sheets
file_path_scheduled = glob.glob(os.path.join(directory_path, "High_Impact_Practices_Scheduled_Learning_AY_*.xlsx"))

if not file_path_scheduled:
    print("No file found with the specified pattern for Scheduled Learning.")
else:
    try:
        # Read HIP and ASL data
        df_hip = pd.read_excel(file_path_scheduled[0], sheet_name='HIP')
        df_asl = pd.read_excel(file_path_scheduled[0], sheet_name='ASL')

        # Clean and merge data as in your code
        df_hip['HEGIS Code'] = df_hip['HEGIS Code'].str.strip()
        df_asl['HEGIS Code'] = df_asl['HEGIS Code'].str.strip()
        df_hip['Home Campus/Teaching Site (Most Recent)'] = df_hip['Home Campus/Teaching Site (Most Recent)'].str.strip()
        df_asl['Home Campus/Teaching Site (Most Recent)'] = df_asl['Home Campus/Teaching Site (Most Recent)'].str.strip()

        # Drop duplicates and merge HIP with ASL
        df_hip = df_hip.drop_duplicates()
        df_asl = df_asl.drop_duplicates()
        df_hip_b = df_hip.drop(columns=['HIP'])

        df_merged = pd.merge(
            df_hip_b,
            df_asl[['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'ASL']],
            on=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)'],
            how='inner'
        )

        # Load and merge Count data
        file_path_directed = glob.glob(os.path.join(directory_path, "High_Impact_Practices_Directed_Service_Learning_AY_*.xlsx"))
        
        if not file_path_directed:
            print("No file found with the specified pattern for Directed Service Learning.")
        else:
            df_count = pd.read_excel(file_path_directed[0], sheet_name='Count')
            df_count['HEGIS Code'] = df_count['HEGIS Code'].str.strip()
            df_count['Home Campus/Teaching Site (Most Recent)'] = df_count['Home Campus/Teaching Site (Most Recent)'].str.strip()
            df_count = df_count.drop_duplicates()

            # Merge ASL data with Count data
            df_final = pd.merge(
                df_merged,
                df_count[['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)', 'Count']],
                on=['HEGIS Code', 'Home Campus/Teaching Site (Most Recent)'],
                how='left'
            )

            # Ensure columns are numeric and calculate sum
            df_final['HIP.1'] = pd.to_numeric(df_final['HIP.1'], errors='coerce').fillna(0)
            df_final['ASL'] = pd.to_numeric(df_final['ASL'], errors='coerce').fillna(0)
            df_final['Count'] = pd.to_numeric(df_final['Count'], errors='coerce').fillna(0)
            df_final['sum'] = df_final[['HIP.1', 'ASL', 'Count']].sum(axis=1)
            df_final['Weighted Sum'] = df_final['sum'] * 0.175

            # Create final rows for HBG, USMGC, and TOTAL for each HEGIS code
            final_rows = []
            for hegis_code in df_final['HEGIS Code'].unique():
                sub_df = df_final[df_final['HEGIS Code'] == hegis_code]

                # Ensure 'HBG' entry
                if 'HBG' in sub_df['Home Campus/Teaching Site (Most Recent)'].values:
                    hbg_row = sub_df[sub_df['Home Campus/Teaching Site (Most Recent)'] == 'HBG']
                else:
                    # If 'HBG' does not exist, add a row with zero values
                    hbg_row = pd.DataFrame([[hegis_code, 'HBG', 0, 0, 0, 0, 0]], columns=df_final.columns)
                final_rows.append(hbg_row)

                # Ensure 'USMGC' entry
                if 'USMGC' in sub_df['Home Campus/Teaching Site (Most Recent)'].values:
                    usmgc_row = sub_df[sub_df['Home Campus/Teaching Site (Most Recent)'] == 'USMGC']
                else:
                    # If 'USMGC' does not exist, add a row with zero values
                    usmgc_row = pd.DataFrame([[hegis_code, 'USMGC', 0, 0, 0, 0, 0]], columns=df_final.columns)
                final_rows.append(usmgc_row)

                # Calculate TOTAL row for the HEGIS code
                total_values = sub_df[['HIP.1', 'ASL', 'Count']].sum()
                total_row = pd.DataFrame({
                    'HEGIS Code': [hegis_code],
                    'Home Campus/Teaching Site (Most Recent)': ['TOTAL'],
                    'HIP.1': [total_values['HIP.1']],
                    'ASL': [total_values['ASL']],
                    'Count': [total_values['Count']],
                    'sum': [total_values.sum()],
                    'Weighted Sum': [total_values.sum() * 0.175]
                })
                final_rows.append(total_row)

            # Combine all rows into a single DataFrame
            final_df = pd.concat(final_rows, ignore_index=True).drop_duplicates()

            # Remove rows where 'HEGIS Code' is missing or blank
            final_df = final_df[final_df['HEGIS Code'].notna() & (final_df['HEGIS Code'] != '')]

            # Specify the parent directory path
            parent_directory = os.path.dirname(directory_path)  # Get the parent directory of your current directory

            # Define the OUTPUT folder path
            output_folder = os.path.join(parent_directory, 'OUTPUT')

            # Ensure the OUTPUT folder exists
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)

            # Set the new file path for saving the Excel file to the OUTPUT folder
            new_file_path = os.path.join(output_folder, "HIP_B.xlsx")

            # Write the final DataFrame to an Excel file in the OUTPUT folder
            final_df.to_excel(new_file_path, sheet_name='Merged Data', index=False)

            # Continue with the other steps (e.g., flattening the data)
            merged_file_path = new_file_path  # Updated path to the OUTPUT folder

            # Load the merged data from the previous step
            df_merged = pd.read_excel(merged_file_path, sheet_name='Merged Data')

            # Pivot the data to reshape it
            df_flattened = df_merged.pivot_table(
                index='HEGIS Code', 
                columns='Home Campus/Teaching Site (Most Recent)', 
                values=['HIP.1', 'ASL', 'Count', 'sum', 'Weighted Sum'], 
                aggfunc='first'  # Since we expect one row per HEGIS code and teaching site
            )

            # Flatten the multi-level columns that result from the pivot
            df_flattened.columns = [' '.join(col).strip() for col in df_flattened.columns.values]

            # Reset the index to make HEGIS Code a column
            df_flattened.reset_index(inplace=True)

            # Write the flattened data to a new sheet in the same file
            with pd.ExcelWriter(merged_file_path, engine='openpyxl', mode='a') as writer:
                df_flattened.to_excel(writer, sheet_name='Flattened Data', index=False)

            #print(f"Flattened data has been written to 'Flattened Data' sheet in '{merged_file_path}'.")

            print(f"File created: '{new_file_path}'")

    except Exception as e:
        print("An error occurred:", e)

print("Process completed successfully for all parts.")