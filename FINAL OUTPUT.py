import pandas as pd
import os
import glob
import openpyxl

######################################################################
# Part 1: Filter HEGIS Codes, Append "STANDARD"
# This pulls from INSTRUCTIONAL_FTE_****, you can change the HEGIS
# codes however you want. 
# Standard is pulled after the specific HEGIS and is placed at the top
######################################################################

# Prompt user for the base directory
directory_path = input("Enter the base directory path: ")
# C:\Users\\The University of Southern Mississippi\IR Office - Documents\W Drive\Userfiles\AKale\Resource Allocation Rubric\AY_**_**\OUTPUT

# List to store DataFrames for HEGIS Codes
dataframes = []

# Define the specific HEGIS Codes
specific_hegis_codes = [
    'ACCOUNT', 'ANTHSOC', 'ARCHTECH', 'ART', 'CELLMOLB', 
    'CHEMBIO', 'CHDFAMSC', 'COASTRES', 'COASTAL', 'COMMSTD', 
    'COMPENG', 'COMPSCI', 'CONSMAN', 'CRIMJUST', 'CURRINST', 
    'DANCE', 'ECOLORGB', 'ECODEVL', 'ECOINTDV', 'EDURESAD', 
    'ENGLISH', 'ENTREPRE', 'FINANCE', 'FORENSIC', 'GENLBUS', 
    'GEO', 'HLTHMKT', 'HISTORY', 'HTM', 'HUMCAPDV', 'HUMANRES', 
    'INDENGTE', 'INFOTECH', 'INTDISC', 'INTERDES', 'INTLDEV', 
    'JOUPRAD', 'KINESIO', 'LEADSHIP', 'LDRADVNP', 'LIBINFSC', 
    'MGT', 'MARINESC', 'MKT', 'MATHD', 'MEDLABSC', 'MERCH', 
    'MUSICAC', 'MUSICED', 'MUSICPER', 'NUTRFOOD', 'OCEANENG', 
    'PHILREL', 'PHYSICS', 'POLISCI', 'POLYMER', 'PRONSGPR', 
    'PSYCH', 'PUBHLTH', 'REHABSC', 'SCIMATED', 'SOCIALWK', 
    'SPECEDU', 'SPEHEASC', 'SPRTMGT', 'THEATRE', 'WLANG'
]

# Construct the file pattern to match Excel files
file_pattern = os.path.join(directory_path, 'INSTRUCTIONAL_FTE_*.xlsx')
print(f"Looking for files matching pattern: {file_pattern}")

# Loop through matching Excel files
matching_files = glob.glob(file_pattern)
print(f"Found files: {matching_files}")

for file_path in matching_files:
    try:
        print(f"Processing file: {file_path}")
        
        # Read the specified sheet from the Excel file
        df = pd.read_excel(file_path, sheet_name='Pivot Table NEW CALC FTE')
        
        # Check if 'HEGIS Code' column exists
        if 'HEGIS Code' in df.columns:
            # Filter DataFrame for specified HEGIS Codes
            filtered_df = df[df['HEGIS Code'].isin(specific_hegis_codes)][['HEGIS Code']]
            
            # Add "STANDARD" to the list
            standard_df = pd.DataFrame({'HEGIS Code': ['STANDARD']})
            final_df = pd.concat([standard_df, filtered_df], ignore_index=True)
            
            # Sort HEGIS Codes by custom order
            final_df['HEGIS Code'] = pd.Categorical(final_df['HEGIS Code'], 
                                                    categories=['STANDARD'] + specific_hegis_codes, 
                                                    ordered=True)
            final_df = final_df.sort_values('HEGIS Code')
            
            # Append to the list of DataFrames
            dataframes.append(final_df)
            
            print(f"Processed DataFrame:\n{final_df}")
        else:
            print(f"'HEGIS Code' column not found in {file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {e}")

# Combine all DataFrames into one
if dataframes:
    combined_df = pd.concat(dataframes, ignore_index=True)
    
    # Write the combined DataFrame to an Excel file
    output_file = os.path.join(directory_path, 'FINAL_OUTPUT.xlsx')
    combined_df.to_excel(output_file, index=False, sheet_name='Filtered HEGIS Codes')
    print(f"Final output saved to: {output_file}")
else:
    print("No valid data to write to the output file.")

##################################################################################################
# Part 2: Filter INSTRUCTIONAL_EFFORT_PART 1 & PART_2

            # MY COLUMN NAMES                      ==           YOUR COLUMN NAMES
    # Total Score:                                 ==   Instructional Effort Part 1 
    # HBG Score:                                   ==   Instructional Effort Part 1 - HBG
    # USMGC Score:                                 ==   Instructional Effort Part 1 - USMGC
    # SCORE:                                       ==   Instructional Effort Part 2 
    # SCORE HBG:                                   ==   Instructional Effort Part 2 - HBG
    # SCORE USMGC:                                 ==   Instructional Effort Part 2 - USMGC
    # Instructional Effort Total Score:            ==   Instructional Effort Total Score
    # Instructional Effort Total Score - HBG:      ==	Instructional Effort Total Score - HBG
    # Instructional Effort Total Score - USMGC:    ==   Instructional Effort Total Score - USMGC
##################################################################################################

# Path to FINAL_OUTPUT.xlsx
final_output_path = os.path.join(directory_path, 'FINAL_OUTPUT.xlsx')
if not os.path.exists(final_output_path):
    print(f"Error: {final_output_path} does not exist.")
    exit()

# Step 1: Load HEGIS Codes from the "Filtered HEGIS Codes" sheet in FINAL_OUTPUT.xlsx
filtered_hegis_df = pd.read_excel(final_output_path, sheet_name='Filtered HEGIS Codes')

# Step 2: Process INSTRUCTIONAL_EFFORT_PART_1
part_1_pattern = os.path.join(directory_path, 'INSTRUCTIONAL_EFFORT_PART_1.xlsx')
part_1_files = glob.glob(part_1_pattern)

if not part_1_files:
    print("No INSTRUCTIONAL_EFFORT_PART_1 file found.")
    exit()

part_1_file_path = part_1_files[0]  # Assuming there's only one matching file
print(f"Processing PART_1 file: {part_1_file_path}")

# Read PART_1 Summary
part_1_df = pd.read_excel(part_1_file_path, sheet_name='Summary Table')

# Extract required columns
part_1_columns = ['HEGIS Code', 'Total Score', 'HBG Score', 'USMGC Score']
if not all(col in part_1_df.columns for col in part_1_columns):
    print(f"Missing columns in PART_1 Summary: {part_1_columns}")
    exit()

part_1_df = part_1_df[part_1_columns]

# Step 3: Process INSTRUCTIONAL_EFFORT_PART_2
part_2_pattern = os.path.join(directory_path, 'INSTRUCTIONAL_EFFORT_PART_2.xlsx')
part_2_files = glob.glob(part_2_pattern)

if not part_2_files:
    print("No INSTRUCTIONAL_EFFORT_PART_2 file found.")
    exit()

part_2_file_path = part_2_files[0]  # Assuming there's only one matching file
print(f"Processing PART_2 file: {part_2_file_path}")

# Read PART_2 Summary
part_2_df = pd.read_excel(part_2_file_path, sheet_name='Summary Table')

# Extract required columns
part_2_columns = ['HEGIS Code', 'SCORE', 'SCORE HBG', 'SCORE USMGC']
if not all(col in part_2_df.columns for col in part_2_columns):
    print(f"Missing columns in PART_2 Summary: {part_2_columns}")
    exit()

part_2_df = part_2_df[part_2_columns]

# Step 4: Merge PART_1 and PART_2 data
merged_effort_df = pd.merge(part_1_df, part_2_df, on='HEGIS Code', how='outer')

# Step 5: Add new columns for instructional effort total scores
merged_effort_df['Instructional Effort Total Score'] = (
    merged_effort_df['Total Score'].fillna(0).astype(float) + 
    merged_effort_df['SCORE'].fillna(0).astype(float)
)

merged_effort_df['Instructional Effort Total Score - HBG'] = (
    merged_effort_df['HBG Score'].fillna(0).astype(float) + 
    merged_effort_df['SCORE HBG'].fillna(0).astype(float)
)

merged_effort_df['Instructional Effort Total Score - USMGC'] = (
    merged_effort_df['USMGC Score'].fillna(0).astype(float) + 
    merged_effort_df['SCORE USMGC'].fillna(0).astype(float)
)

# Step 6: Calculate the "Standard" row (max of all the columns)
columns_to_consider = [
    'Total Score', 'HBG Score', 'USMGC Score', 'SCORE', 'SCORE HBG', 
    'SCORE USMGC', 'Instructional Effort Total Score', 
    'Instructional Effort Total Score - HBG', 'Instructional Effort Total Score - USMGC'
]

# Calculate the max for each column
standard_values = merged_effort_df[columns_to_consider].max()

# Create a "Standard" row with the calculated max values
standard_row = pd.DataFrame(standard_values).transpose()
standard_row['HEGIS Code'] = 'STANDARD'

# Step 7: Reorganize the DataFrame to put the "Standard" row at the top
final_effort_df = pd.concat([standard_row, merged_effort_df], ignore_index=True)

# Step 8: Merge the new data into the "Filtered HEGIS Codes" DataFrame
updated_hegis_df = pd.merge(filtered_hegis_df, final_effort_df, on='HEGIS Code', how='left')

# Fill NaN values with 0.00 and ensure all columns are numeric before formatting
updated_hegis_df.fillna(0.00, inplace=True)  # Replace NaN values with 0.00

# Convert all relevant columns to float to ensure consistency
columns_to_format = updated_hegis_df.columns.difference(['HEGIS Code'])

# Round all the numeric values to 2 decimal places
for col in columns_to_format:
    updated_hegis_df[col] = updated_hegis_df[col].round(2)

# If you need to format as string with two decimals
for col in columns_to_format:
    updated_hegis_df[col] = updated_hegis_df[col].apply(lambda x: f"{x:.2f}")

# This ensures all numeric columns are rounded and formatted to two decimal places as strings.

# Step 10: Write the updated DataFrame back to FINAL_OUTPUT.xlsx
print("Updating the 'Filtered HEGIS Codes' sheet in FINAL_OUTPUT.xlsx...")
with pd.ExcelWriter(final_output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    updated_hegis_df.to_excel(writer, index=False, sheet_name='Filtered HEGIS Codes')

print(f"'Filtered HEGIS Codes' sheet successfully updated in FINAL_OUTPUT.xlsx.")


##################################################################################################
# Part 3: Filter SUCCES_PART_1 & PART_2

            # MY COLUMN NAMES                      ==           YOUR COLUMN NAMES
    # Adjusted_Ratio:	                           ==   Success Part 1
    # Adjusted_Ratio_HBG:                          ==   Success Part 1 - HBG
    # Adjusted_Ratio_USMGC:                        ==	Success Part 1 - USMGC
    # Total Scores:	                               ==   Success Part 2
    # Hattiesburg Score:                           == 	Success Part 2 - HBG
    # USM Gulf Coast Score:                        ==	Success Part 2 - USMGC
    # Success Total Score:                         ==   Success Total Score
    # Success Total Score - HBG:                   ==   Success Total Score - HBG
    # Success Total Score - USMGC:                 ==   Success Total Score - USMGC

##################################################################################################

import os
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Path to FINAL_OUTPUT.xlsx
final_output_path = os.path.join(directory_path, 'FINAL_OUTPUT.xlsx')
if not os.path.exists(final_output_path):
    print(f"Error: {final_output_path} does not exist.")
    exit()

# Step 1: Load HEGIS Codes from the "Filtered HEGIS Codes" sheet in FINAL_OUTPUT.xlsx
filtered_hegis_df = pd.read_excel(final_output_path, sheet_name='Filtered HEGIS Codes')

# Step 2: Process SUCCESS_PART_1
part_1_pattern = os.path.join(directory_path, 'SUCCESS_PART_1.xlsx')
part_1_files = glob.glob(part_1_pattern)

if not part_1_files:
    print("No SUCCESS_PART_1 file found.")
    exit()

part_1_file_path = part_1_files[0]  # Assuming there's only one matching file
print(f"Processing PART_1 file: {part_1_file_path}")

# Read PART_1 Summary
part_1_df = pd.read_excel(part_1_file_path, sheet_name='Summary')

# Extract required columns
part_1_columns = ['HEGIS Code', 'Adjusted_Ratio', 'Adjusted_Ratio_HBG', 'Adjusted_Ratio_USMGC']
if not all(col in part_1_df.columns for col in part_1_columns):
    print(f"Missing columns in PART_1 Summary: {part_1_columns}")
    exit()

part_1_df = part_1_df[part_1_columns]

# Step 3: Process SUCCESS_PART_2
part_2_pattern = os.path.join(directory_path, 'SUCCESS_PART_2.xlsx')
part_2_files = glob.glob(part_2_pattern)

if not part_2_files:
    print("No SUCCESS_PART_2 file found.")
    exit()

part_2_file_path = part_2_files[0]  # Assuming there's only one matching file
print(f"Processing PART_2 file: {part_2_file_path}")

# Read PART_2 Summary
part_2_df = pd.read_excel(part_2_file_path, sheet_name='Summary')

# Extract required columns
part_2_columns = ['HEGIS Code', 'Total Scores', 'Hattiesburg Score', 'USM Gulf Coast Score']
if not all(col in part_2_df.columns for col in part_2_columns):
    print(f"Missing columns in PART_2 Summary: {part_2_columns}")
    exit()

part_2_df = part_2_df[part_2_columns]

# Step 4: Merge PART_1 and PART_2 data
merged_success_df = pd.merge(part_1_df, part_2_df, on='HEGIS Code', how='outer')

# Step 5: Add new columns for success total scores
merged_success_df['Success Total Score'] = (
    merged_success_df['Adjusted_Ratio'].fillna(0).astype(float) +
    merged_success_df['Total Scores'].fillna(0).astype(float)
)

merged_success_df['Success Total Score - HBG'] = (
    merged_success_df['Adjusted_Ratio_HBG'].fillna(0).astype(float) +
    merged_success_df['Hattiesburg Score'].fillna(0).astype(float)
)

merged_success_df['Success Total Score - USMGC'] = (
    merged_success_df['Adjusted_Ratio_USMGC'].fillna(0).astype(float) +
    merged_success_df['USM Gulf Coast Score'].fillna(0).astype(float)
)

# Step 6: Calculate the "Standard" row (max of all the columns)
columns_to_consider = [
    'Adjusted_Ratio', 'Adjusted_Ratio_HBG', 'Adjusted_Ratio_USMGC', 
    'Total Scores', 'Hattiesburg Score', 'USM Gulf Coast Score',
    'Success Total Score','Success Total Score - HBG', 'Success Total Score - USMGC'
]

# Calculate the max for each column
standard_values = merged_success_df[columns_to_consider].max()

# Create a "Standard" row with the calculated max values
standard_row = pd.DataFrame(standard_values).transpose()
standard_row['HEGIS Code'] = 'STANDARD'

# Step 7: Reorganize the DataFrame to put the "Standard" row at the top
final_success_df = pd.concat([standard_row, merged_success_df], ignore_index=True)

# Step 8: Merge the new data into the "Filtered HEGIS Codes" DataFrame
updated_hegis_df = pd.merge(filtered_hegis_df, final_success_df, on='HEGIS Code', how='left')

# Fill NaN values with 0.00 and ensure all columns are numeric before formatting
updated_hegis_df.fillna(0.00, inplace=True)  # Replace NaN values with 0.00

# Convert all relevant columns to float to ensure consistency
columns_to_format = updated_hegis_df.columns.difference(['HEGIS Code'])

# Format all the numeric values to 2 decimal places as strings (not rounding)
for col in columns_to_format:
    updated_hegis_df[col] = updated_hegis_df[col].apply(lambda x: f"{x:.2f}")

# Step 10: Write the updated DataFrame back to FINAL_OUTPUT.xlsx with openpyxl formatting
print("Updating the 'Filtered HEGIS Codes' sheet in FINAL_OUTPUT.xlsx...")
with pd.ExcelWriter(final_output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    updated_hegis_df.to_excel(writer, index=False, sheet_name='Filtered HEGIS Codes')

    # Apply formatting to the Excel file using openpyxl
    workbook = writer.book
    sheet = workbook['Filtered HEGIS Codes']

    # Define a number format for two decimal places
    number_format = '#,##0.00'

    # Apply the number format to all the cells in the columns with numeric values
    for col_idx, col in enumerate(updated_hegis_df.columns, start=1):
        for row in range(2, len(updated_hegis_df) + 2):  # Start from row 2 to skip the header
            cell = sheet.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format

print(f"'Filtered HEGIS Codes' sheet successfully updated in FINAL_OUTPUT.xlsx.")

##################################################################################################
# Part 4: Filter FS_A_updated & HIP_B

          # MY COLUMN NAMES                        ==           YOUR COLUMN NAMES
    # total:                                       ==   Success Total Score
    # hbg:	                                       ==   Success Total Score - HBG
    # usmgc:                                       ==   Success Total Score - USMGC
    # Weighted Sum TOTAL:                          ==   Engagment Part 1 
    # Weighted Sum HBG:	                           ==   Engagement Part 1 - HBG
    # Weighted Sum USMGC:                          ==	Engagement Part 1- USMGC
    # Engagement Total Score:                      ==	Engagement Part 2
    # Engagement Total HBG:	                       ==   Engagement Part 2 - HBG
    # Engagement Total USMGC:                      ==   Engagement Part 2 - USMGC
##################################################################################################

# Path to FINAL_OUTPUT.xlsx
final_output_path = os.path.join(directory_path, 'FINAL_OUTPUT.xlsx')
if not os.path.exists(final_output_path):
    print(f"Error: {final_output_path} does not exist.")
    exit()

# Step 1: Load HEGIS Codes from the "Filtered HEGIS Codes" sheet in FINAL_OUTPUT.xlsx
filtered_hegis_df = pd.read_excel(final_output_path, sheet_name='Filtered HEGIS Codes')

# Step 2: Process FS_A (Flattened Data sheet)
fs_a_pattern = os.path.join(directory_path, 'FS_A_updated.xlsx')
fs_a_files = glob.glob(fs_a_pattern)

if not fs_a_files:
    print("No FS_A file found.")
    exit()

fs_a_file_path = fs_a_files[0]  # Assuming there's only one matching file
print(f"Processing FS_A file: {fs_a_file_path}")

# Read FS_A Flattened Data sheet
fs_a_df = pd.read_excel(fs_a_file_path, sheet_name='Flattened Data')

# Extract required columns
fs_a_columns = ['HEGIS Code', 'hbg', 'total', 'usmgc']
if not all(col in fs_a_df.columns for col in fs_a_columns):
    print(f"Missing columns in FS_A: {fs_a_columns}")
    exit()

fs_a_df = fs_a_df[fs_a_columns]

# Step 3: Process HIP_B (Flattened Data sheet)
hip_b_pattern = os.path.join(directory_path, 'HIP_B.xlsx')
hip_b_files = glob.glob(hip_b_pattern)

if not hip_b_files:
    print("No HIP_B file found.")
    exit()

hip_b_file_path = hip_b_files[0]  # Assuming there's only one matching file
print(f"Processing HIP_B file: {hip_b_file_path}")

# Read HIP_B Flattened Data sheet
hip_b_df = pd.read_excel(hip_b_file_path, sheet_name='Flattened Data')

# Extract required columns
hip_b_columns = ['HEGIS Code', 'Weighted Sum HBG', 'Weighted Sum TOTAL', 'Weighted Sum USMGC']
if not all(col in hip_b_df.columns for col in hip_b_columns):
    print(f"Missing columns in HIP_B: {hip_b_columns}")
    exit()

hip_b_df = hip_b_df[hip_b_columns]

# Step 4: Merge FS_A and HIP_B data
merged_fs_hip_df = pd.merge(fs_a_df, hip_b_df, on='HEGIS Code', how='outer')

# Step 5: Add new columns for EN scores
merged_fs_hip_df['Engagement Total Score'] = (
    merged_fs_hip_df['total'].fillna(0).astype(float) + 
    merged_fs_hip_df['Weighted Sum TOTAL'].fillna(0).astype(float)
)

merged_fs_hip_df['Engagement Total HBG'] = (
    merged_fs_hip_df['hbg'].fillna(0).astype(float) + 
    merged_fs_hip_df['Weighted Sum HBG'].fillna(0).astype(float)
)

merged_fs_hip_df['Engagement Total USMGC'] = (
    merged_fs_hip_df['usmgc'].fillna(0).astype(float) + 
    merged_fs_hip_df['Weighted Sum USMGC'].fillna(0).astype(float)
)

# Step 6: Calculate the "Standard" row (max of all the columns)
columns_to_consider_en = [
    'total', 'hbg', 'usmgc', 'Weighted Sum TOTAL', 'Weighted Sum HBG', 
    'Weighted Sum USMGC', 'Engagement Total Score', 'Engagement Total HBG', 
    'Engagement Total USMGC'
]

# Calculate the max for each column
standard_values_en = merged_fs_hip_df[columns_to_consider_en].max()

# Create a "Standard" row with the calculated max values
standard_row_en = pd.DataFrame(standard_values_en).transpose()
standard_row_en['HEGIS Code'] = 'STANDARD'

# Step 7: Reorganize the DataFrame to put the "Standard" row at the top
final_en_df = pd.concat([standard_row_en, merged_fs_hip_df], ignore_index=True)

# Step 8: Merge the new data into the "Filtered HEGIS Codes" DataFrame
updated_hegis_df = pd.merge(filtered_hegis_df, final_en_df, on='HEGIS Code', how='left')

# Step 9: Fill NaN values with 0.00 and format scores to 2 decimal places
updated_hegis_df.fillna(0.00, inplace=True)  # Replace NaN with 0.00

# Ensure the columns to format are numeric (float) before applying the formatting
columns_to_format_en = final_en_df.columns.difference(['HEGIS Code'])

# Convert all relevant columns to float first
for col in columns_to_format_en:
    updated_hegis_df[col] = updated_hegis_df[col].astype(float)  # Ensure column is numeric

# Now, apply 2 decimal formatting
for col in columns_to_format_en:
    updated_hegis_df[col] = updated_hegis_df[col].apply(lambda x: f"{x:.2f}")  # Format to 2 decimal places

# Step 10: Write the updated DataFrame back to FINAL_OUTPUT.xlsx
print("Updating the 'Filtered HEGIS Codes' sheet in FINAL_OUTPUT.xlsx...")
with pd.ExcelWriter(final_output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    updated_hegis_df.to_excel(writer, index=False, sheet_name='Filtered HEGIS Codes')

print(f"'Filtered HEGIS Codes' sheet successfully updated in FINAL_OUTPUT.xlsx.")

##################################################################################################
# Step 5: Add new columns for rubric scores
# Column naming is the same
    # Rubric Total Score
    # Rubric Total Score - HBG	
    # Rubric Total Score - USMGC	
    # Rubric Standardized Score	
    # Rubric Standardized Score - HBG	
    # Rubric Standardized Score - USMGC
##################################################################################################

import pandas as pd
import openpyxl

# Load the workbook and the first sheet
wb = openpyxl.load_workbook(final_output_path)
sheet = wb['Filtered HEGIS Codes']

# Step 1: Read the existing data into a DataFrame
data = sheet.values
columns = next(data)  # First row as column names
existing_df = pd.DataFrame(data, columns=columns)

# Step 2: Calculate the new columns (Rubric Total Scores and Standardized Scores)
required_columns = [
    'Instructional Effort Total Score', 'Success Total Score', 'Engagement Total Score',
    'Instructional Effort Total Score - HBG', 'Success Total Score - HBG', 'Engagement Total HBG',
    'Instructional Effort Total Score - USMGC', 'Success Total Score - USMGC', 'Engagement Total USMGC'
]

# Add missing columns with default values (if not present)
for col in required_columns:
    if col not in existing_df.columns:
        existing_df[col] = 0.00  # Default to zero

# Calculate Rubric Total Scores
existing_df['Rubric Total Score'] = (
    existing_df['Instructional Effort Total Score'].fillna(0).astype(float) +
    existing_df['Success Total Score'].fillna(0).astype(float) +
    existing_df['Engagement Total Score'].fillna(0).astype(float)
)

existing_df['Rubric Total Score - HBG'] = (
    existing_df['Instructional Effort Total Score - HBG'].fillna(0).astype(float) +
    existing_df['Success Total Score - HBG'].fillna(0).astype(float) +
    existing_df['Engagement Total HBG'].fillna(0).astype(float)
)

existing_df['Rubric Total Score - USMGC'] = (
    existing_df['Instructional Effort Total Score - USMGC'].fillna(0).astype(float) +
    existing_df['Success Total Score - USMGC'].fillna(0).astype(float) +
    existing_df['Engagement Total USMGC'].fillna(0).astype(float)
)

# Step 3: Calculate Standardized Scores (as percentages)
existing_df['Rubric Standardized Score'] = (
    (existing_df['Rubric Total Score'].fillna(0).astype(float) / 278) * 100
)

existing_df['Rubric Standardized Score - HBG'] = (
    (existing_df['Rubric Total Score - HBG'].fillna(0).astype(float) / 259.72) * 100
)

existing_df['Rubric Standardized Score - USMGC'] = (
    (existing_df['Rubric Total Score - USMGC'].fillna(0).astype(float) / 140.68) * 100
)

# Step 3.5: Set specific row values to 100 for 'STANDARD'
standard_row_condition = existing_df['HEGIS Code'] == 'STANDARD'  # Adjust 'HEGIS Code' to match your column
standardized_columns = [
    'Rubric Standardized Score',
    'Rubric Standardized Score - HBG',
    'Rubric Standardized Score - USMGC'
]

existing_df.loc[standard_row_condition, standardized_columns] = 100.00

# Step 4: Format columns to two decimal places
decimal_columns = [
    'Rubric Total Score', 'Rubric Total Score - HBG', 'Rubric Total Score - USMGC'
]

# Format these columns to two decimal places
for col in decimal_columns:
    existing_df[col] = existing_df[col].round(2)

# Format specific columns with percentages and add % symbol
percent_columns = [
    'Rubric Standardized Score', 'Rubric Standardized Score - HBG', 'Rubric Standardized Score - USMGC'
]

# Format percentage columns to two decimal places and add % symbol
for col in percent_columns:
    existing_df[col] = existing_df[col].round(2).astype(str) + "%"


# Step 5: Create a new sheet and copy all the existing data over
new_sheet_name = "Updated HEGIS Codes"
if new_sheet_name in wb.sheetnames:
    del wb[new_sheet_name]  # Remove the existing sheet if it exists

# Create a new sheet in the workbook
new_sheet = wb.create_sheet(new_sheet_name)

# Step 6: Write the column headers (first row)
for col_num, col_name in enumerate(existing_df.columns, 1):
    new_sheet.cell(row=1, column=col_num, value=col_name)

# Step 7: Write the updated DataFrame to the new sheet starting from the second row
for r_idx, row in enumerate(existing_df.itertuples(index=False), 2):  # Set index=False to avoid the index being written
    for c_idx, value in enumerate(row, 1):  # Start from column 1
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(final_output_path)

print(f"Successfully updated with new sheet '{new_sheet_name}' in {final_output_path}.")